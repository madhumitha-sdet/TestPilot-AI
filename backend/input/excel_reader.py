from pathlib import Path

from openpyxl import load_workbook

from models.test_case import TestCase


class ExcelReader:
    """Reads test case data from an Excel file and converts rows into TestCase objects."""

    def __init__(self, file_path: str) -> None:
        """
        Initialize the reader with the path to the Excel file.

        Args:
            file_path: Path to the .xlsx file containing test case data.
        """
        self.file_path = Path(file_path)

    def read_test_cases(self) -> list[TestCase]:
        """
        Read all rows from the Excel file and convert them into TestCase objects.

        Returns:
            A list of TestCase objects parsed from the sheet.

        Raises:
            FileNotFoundError: If the Excel file does not exist.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        workbook = load_workbook(filename=self.file_path, read_only=True, data_only=True)
        sheet = workbook.active

        test_cases: list[TestCase] = []

        rows = sheet.iter_rows(min_row=2, values_only=True)  # skip header row
        for row in rows:
            if row is None or all(cell is None for cell in row):
                continue  # skip completely empty rows

            test_case_id, title, description, expected_result, test_data = (
                row[0], row[1], row[2], row[3], row[4]
            )

            test_cases.append(
                TestCase(
                    test_case_id=str(test_case_id) if test_case_id is not None else "",
                    title=str(title) if title is not None else "",
                    description=str(description) if description is not None else "",
                    expected_result=str(expected_result) if expected_result is not None else "",
                    test_data=str(test_data) if test_data is not None else "",
                )
            )

        workbook.close()
        return test_cases